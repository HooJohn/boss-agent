"""Utility function for reading file content with support for various formats."""

import os
import mammoth
import openpyxl
from boss_agent.tools.advanced_tools.pdf_tool import PdfTextExtractTool
from boss_agent.utils import WorkspaceManager

def read_file_content(file_path: str, workspace_manager: WorkspaceManager) -> str:
    """
    Read the content of a file, with support for pdf, docx, and xlsx.
    
    Args:
        file_path: The absolute path to the file.
        workspace_manager: The workspace manager instance.

    Returns:
        The content of the file as a string, or an error message.
    """
    try:
        if not os.path.exists(file_path):
            return f"Error reading file: File not found at {file_path}"

        if file_path.endswith(".pdf"):
            pdf_extractor = PdfTextExtractTool(workspace_manager)
            return pdf_extractor.run_impl({"path": file_path}).result
        elif file_path.endswith(".docx"):
            with open(file_path, "rb") as docx_file:
                result = mammoth.convert_to_html(docx_file)
                return result.value
        elif file_path.endswith(".xlsx"):
            workbook = openpyxl.load_workbook(file_path)
            content = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content.append(f"Sheet: {sheet_name}\\n")
                for row in sheet.iter_rows(values_only=True):
                    row_content = [str(cell) if cell is not None else "" for cell in row]
                    content.append(", ".join(row_content))
            return "\\n".join(content)
        elif file_path.endswith((".txt", ".md", ".html", ".csv", ".json")):
            with open(file_path, "r", encoding="utf-8") as f:
                return f.read()
        else:
            # For other file types, we don't attempt to read the content
            return f"Unsupported file type: {os.path.basename(file_path)}"
    except Exception as e:
        return f"Error reading file {file_path}: {e}"
