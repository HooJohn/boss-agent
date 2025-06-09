"""Tool for reading the content of a specific file."""

import os
import mammoth
import openpyxl
import pandas as pd
from typing import Any, Optional, List
from boss_agent.tools.base import LLMTool, ToolImplOutput
from boss_agent.utils import WorkspaceManager

class ReadFileTool(LLMTool):
    name = "read_file"
    description = "Reads the entire content of a single specified file from the knowledge base."

    input_schema = {
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": "The exact path to the file, relative to the knowledge base root.",
            }
        },
        "required": ["path"],
    }

    def __init__(self, workspace_manager: WorkspaceManager):
        super().__init__()
        self.workspace_manager = workspace_manager

    def _read_file_content(self, file_path: str) -> str:
        """Read the content of a file, with support for various formats."""
        try:
            if file_path.endswith(".docx"):
                with open(file_path, "rb") as docx_file:
                    result = mammoth.convert_to_html(docx_file)
                    return result.value
            elif file_path.endswith(".xlsx"):
                workbook = openpyxl.load_workbook(file_path)
                content = []
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    content.append(f"Sheet: {sheet_name}\\n")
                    for row in sheet.iter_rows(values_only=True):
                        row_content = [str(cell) if cell is not None else "" for cell in row]
                        content.append(", ".join(row_content))
                return "\\n".join(content)
            elif file_path.endswith((".txt", ".md", ".html", ".csv", ".json")):
                with open(file_path, "r", encoding="utf-8") as f:
                    return f.read()
            else:
                return f"Error: Unsupported file type for reading: {os.path.basename(file_path)}"
        except Exception as e:
            return f"Error reading file {file_path}: {e}"

    def run_impl(
        self,
        tool_input: dict[str, Any],
        message_history: Optional[List[dict]] = None,
    ) -> ToolImplOutput:
        file_path_str = tool_input.get("path")
        if not file_path_str:
            return ToolImplOutput("", "Error: 'path' is a required parameter.")

        # --- Path Sanitization ---
        safe_path = os.path.normpath(os.path.join('/', file_path_str)).lstrip('/\\')
        if ".." in safe_path.split(os.sep):
            return ToolImplOutput("", "Error: Directory traversal is not allowed.")

        # Prioritize session path, then fall back to knowledge base
        full_path = os.path.join(self.workspace_manager.session_workspace, safe_path)
        if not os.path.exists(full_path):
            full_path = os.path.join(self.workspace_manager.root, safe_path)

        if not os.path.isfile(full_path):
            return ToolImplOutput("", f"Error: File not found at '{file_path_str}'.")

        content = self._read_file_content(full_path)
        return ToolImplOutput(content, f"Successfully read content from '{file_path_str}'.")
